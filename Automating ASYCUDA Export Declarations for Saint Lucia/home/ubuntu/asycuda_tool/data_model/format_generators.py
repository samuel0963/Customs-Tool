"""
Format generators for ASYCUDA export declarations.

This module provides classes to generate different export formats required by ASYCUDA:
- XML format
- Pipe-delimited text (.txt)
- Structured Excel
- PDF-style declaration

Each generator takes a Declaration object and produces the corresponding output format.
"""

from typing import Dict, List, Any, Optional, BinaryIO, TextIO
import xml.dom.minidom as minidom
import xml.etree.ElementTree as ET
import csv
import io
import datetime
from .asycuda_data_model import Declaration, Item


class FormatGeneratorError(Exception):
    """Exception raised for format generation errors."""
    pass


class XmlGenerator:
    """
    Generates XML format for ASYCUDA import.
    """
    
    def generate(self, declaration: Declaration, output_file: TextIO) -> None:
        """
        Generate XML representation of the declaration.
        
        Args:
            declaration: The Declaration object to convert
            output_file: File-like object to write XML to
            
        Returns:
            None
        """
        # Create root element
        root = ET.Element("AsycudaDeclaration")
        root.set("version", "1.0")
        
        # Add header information
        header = ET.SubElement(root, "Header")
        ET.SubElement(header, "RegistrationNumber").text = declaration.registration_number
        ET.SubElement(header, "DeclarationType").text = declaration.declaration_type
        ET.SubElement(header, "CustomsOffice").text = declaration.customs_office
        ET.SubElement(header, "Date").text = declaration.date.strftime("%d/%m/%Y")
        
        # Add exporter information
        exporter = ET.SubElement(root, "Exporter")
        ET.SubElement(exporter, "ID").text = declaration.exporter.id
        ET.SubElement(exporter, "Name").text = declaration.exporter.name
        ET.SubElement(exporter, "AddressLine1").text = declaration.exporter.address_line1
        if declaration.exporter.address_line2:
            ET.SubElement(exporter, "AddressLine2").text = declaration.exporter.address_line2
        ET.SubElement(exporter, "City").text = declaration.exporter.city
        ET.SubElement(exporter, "Country").text = declaration.exporter.country
        
        # Add declarant information
        declarant = ET.SubElement(root, "Declarant")
        ET.SubElement(declarant, "ID").text = declaration.declarant.id
        ET.SubElement(declarant, "Name").text = declaration.declarant.name
        ET.SubElement(declarant, "AddressLine1").text = declaration.declarant.address_line1
        if declaration.declarant.address_line2:
            ET.SubElement(declarant, "AddressLine2").text = declaration.declarant.address_line2
        ET.SubElement(declarant, "City").text = declaration.declarant.city
        ET.SubElement(declarant, "Country").text = declaration.declarant.country
        
        # Add declaration details
        details = ET.SubElement(root, "DeclarationDetails")
        ET.SubElement(details, "GeneralProcedureCode").text = declaration.general_procedure_code
        ET.SubElement(details, "ExtendedProcedureCode").text = declaration.extended_procedure_code
        ET.SubElement(details, "CountryOfDestination").text = declaration.country_of_destination
        ET.SubElement(details, "ModeOfTransport").text = declaration.mode_of_transport
        ET.SubElement(details, "OfficeOfEntryExit").text = declaration.office_of_entry_exit
        ET.SubElement(details, "CurrencyCode").text = declaration.currency_code
        ET.SubElement(details, "ExchangeRate").text = str(declaration.exchange_rate)
        ET.SubElement(details, "TotalPackages").text = str(declaration.total_packages)
        ET.SubElement(details, "CommercialReference").text = declaration.commercial_reference
        
        if declaration.valuation_method:
            ET.SubElement(details, "ValuationMethod").text = declaration.valuation_method
        if declaration.delivery_terms:
            ET.SubElement(details, "DeliveryTerms").text = declaration.delivery_terms
        if declaration.place_of_loading:
            ET.SubElement(details, "PlaceOfLoading").text = declaration.place_of_loading
        if declaration.manifest_reference:
            ET.SubElement(details, "ManifestReference").text = declaration.manifest_reference
        if declaration.warehouse_identification:
            ET.SubElement(details, "WarehouseIdentification").text = declaration.warehouse_identification
        if declaration.declarant_signature:
            ET.SubElement(details, "DeclarantSignature").text = declaration.declarant_signature
        
        # Add items
        items_element = ET.SubElement(root, "Items")
        for item in declaration.items:
            item_element = ET.SubElement(items_element, "Item")
            ET.SubElement(item_element, "ItemNumber").text = str(item.item_number)
            ET.SubElement(item_element, "HSCode").text = item.hs_code
            ET.SubElement(item_element, "Description").text = item.description
            ET.SubElement(item_element, "CountryOfOrigin").text = item.country_of_origin
            ET.SubElement(item_element, "GrossWeight").text = str(item.gross_weight)
            ET.SubElement(item_element, "NetWeight").text = str(item.net_weight)
            ET.SubElement(item_element, "StatisticalUnit").text = item.statistical_unit
            ET.SubElement(item_element, "Quantity").text = str(item.quantity)
            ET.SubElement(item_element, "CustomsValue").text = str(item.customs_value)
            ET.SubElement(item_element, "PackageType").text = item.package_type
            ET.SubElement(item_element, "PackageCount").text = str(item.package_count)
            ET.SubElement(item_element, "MarksAndNumbers").text = item.marks_and_numbers
            if item.previous_document:
                ET.SubElement(item_element, "PreviousDocument").text = item.previous_document
        
        # Convert to pretty XML and write to file
        xml_string = ET.tostring(root, encoding="utf-8")
        pretty_xml = minidom.parseString(xml_string).toprettyxml(indent="  ")
        output_file.write(pretty_xml)


class PipeDelimitedGenerator:
    """
    Generates pipe-delimited text format for ASYCUDA import.
    """
    
    def generate(self, declaration: Declaration, output_file: TextIO) -> None:
        """
        Generate pipe-delimited representation of the declaration.
        
        Args:
            declaration: The Declaration object to convert
            output_file: File-like object to write to
            
        Returns:
            None
        """
        # Create CSV writer with pipe delimiter
        writer = csv.writer(output_file, delimiter='|', quoting=csv.QUOTE_MINIMAL)
        
        # Write header row
        writer.writerow([
            "H",  # Header indicator
            declaration.registration_number,
            declaration.declaration_type,
            declaration.customs_office,
            declaration.date.strftime("%d/%m/%Y"),
            declaration.exporter.id,
            declaration.exporter.name,
            declaration.exporter.address_line1,
            declaration.exporter.city,
            declaration.exporter.country,
            declaration.declarant.id,
            declaration.declarant.name,
            declaration.general_procedure_code,
            declaration.extended_procedure_code,
            declaration.country_of_destination,
            declaration.mode_of_transport,
            declaration.office_of_entry_exit,
            declaration.currency_code,
            str(declaration.exchange_rate),
            str(declaration.total_packages),
            declaration.commercial_reference,
            declaration.valuation_method or "",
            declaration.delivery_terms or "",
            declaration.place_of_loading or "",
            declaration.manifest_reference or "",
            declaration.warehouse_identification or "",
            declaration.declarant_signature or ""
        ])
        
        # Write item rows
        for item in declaration.items:
            writer.writerow([
                "I",  # Item indicator
                str(item.item_number),
                item.hs_code,
                item.description,
                item.country_of_origin,
                str(item.gross_weight),
                str(item.net_weight),
                item.statistical_unit,
                str(item.quantity),
                str(item.customs_value),
                item.package_type,
                str(item.package_count),
                item.marks_and_numbers,
                item.previous_document or ""
            ])


class ExcelGenerator:
    """
    Generates structured Excel format for ASYCUDA data review.
    """
    
    def generate(self, declaration: Declaration, output_file: BinaryIO) -> None:
        """
        Generate Excel representation of the declaration.
        
        Args:
            declaration: The Declaration object to convert
            output_file: File-like object to write Excel data to
            
        Returns:
            None
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise FormatGeneratorError("pandas and openpyxl are required for Excel generation")
        
        # Create workbook and sheets
        wb = Workbook()
        
        # Create declaration sheet
        declaration_sheet = wb.active
        declaration_sheet.title = "Declaration"
        
        # Add header information
        declaration_sheet['A1'] = "ASYCUDA Export Declaration"
        declaration_sheet['A1'].font = Font(bold=True, size=14)
        declaration_sheet.merge_cells('A1:D1')
        
        # Add declaration details
        declaration_data = [
            ["Registration Number", declaration.registration_number],
            ["Declaration Type", declaration.declaration_type],
            ["Customs Office", declaration.customs_office],
            ["Date", declaration.date.strftime("%d/%m/%Y")],
            ["", ""],
            ["Exporter ID", declaration.exporter.id],
            ["Exporter Name", declaration.exporter.name],
            ["Exporter Address", declaration.exporter.address_line1],
            ["Exporter City", declaration.exporter.city],
            ["Exporter Country", declaration.exporter.country],
            ["", ""],
            ["Declarant ID", declaration.declarant.id],
            ["Declarant Name", declaration.declarant.name],
            ["", ""],
            ["General Procedure Code", declaration.general_procedure_code],
            ["Extended Procedure Code", declaration.extended_procedure_code],
            ["Country of Destination", declaration.country_of_destination],
            ["Mode of Transport", declaration.mode_of_transport],
            ["Office of Entry/Exit", declaration.office_of_entry_exit],
            ["Currency Code", declaration.currency_code],
            ["Exchange Rate", str(declaration.exchange_rate)],
            ["Total Packages", str(declaration.total_packages)],
            ["Commercial Reference", declaration.commercial_reference],
            ["Valuation Method", declaration.valuation_method or ""],
            ["Delivery Terms", declaration.delivery_terms or ""],
            ["Warehouse Identification", declaration.warehouse_identification or ""]
        ]
        
        for row_idx, (label, value) in enumerate(declaration_data, start=3):
            declaration_sheet[f'A{row_idx}'] = label
            declaration_sheet[f'B{row_idx}'] = value
            if label:
                declaration_sheet[f'A{row_idx}'].font = Font(bold=True)
        
        # Create items sheet
        items_sheet = wb.create_sheet(title="Items")
        
        # Create DataFrame for items
        items_data = []
        for item in declaration.items:
            items_data.append({
                "Item #": item.item_number,
                "HS Code": item.hs_code,
                "Description": item.description,
                "Origin": item.country_of_origin,
                "Gross Weight": item.gross_weight,
                "Net Weight": item.net_weight,
                "Unit": item.statistical_unit,
                "Quantity": item.quantity,
                "Value": item.customs_value,
                "Package Type": item.package_type,
                "Packages": item.package_count,
                "Marks": item.marks_and_numbers,
                "Previous Doc": item.previous_document or ""
            })
        
        if items_data:
            df = pd.DataFrame(items_data)
            
            # Add header row
            items_sheet['A1'] = "Declaration Items"
            items_sheet['A1'].font = Font(bold=True, size=14)
            items_sheet.merge_cells('A1:F1')
            
            # Add DataFrame rows starting at row 3
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
                for c_idx, value in enumerate(row, start=1):
                    items_sheet.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 3:  # Header row
                        items_sheet.cell(row=r_idx, column=c_idx).font = Font(bold=True)
                        items_sheet.cell(row=r_idx, column=c_idx).fill = PatternFill(
                            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
                        )
        
        # Create summary sheet
        summary_sheet = wb.create_sheet(title="Summary")
        
        # Calculate totals
        totals = declaration.calculate_totals()
        
        # Add summary data
        summary_sheet['A1'] = "Declaration Summary"
        summary_sheet['A1'].font = Font(bold=True, size=14)
        summary_sheet.merge_cells('A1:C1')
        
        summary_data = [
            ["Total Items", str(totals['total_items'])],
            ["Total Packages", str(totals['total_packages'])],
            ["Total Gross Weight", f"{totals['total_gross_weight']:.2f} kg"],
            ["Total Net Weight", f"{totals['total_net_weight']:.2f} kg"],
            ["Total Value", f"{totals['total_value']:.2f} {declaration.currency_code}"]
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, start=3):
            summary_sheet[f'A{row_idx}'] = label
            summary_sheet[f'B{row_idx}'] = value
            summary_sheet[f'A{row_idx}'].font = Font(bold=True)
        
        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(output_file)


class PDFGenerator:
    """
    Generates PDF-style declaration similar to ASYCUDA's interface.
    """
    
    def generate(self, declaration: Declaration, output_file: BinaryIO) -> None:
        """
        Generate PDF representation of the declaration.
        
        Args:
            declaration: The Declaration object to convert
            output_file: File-like object to write PDF data to
            
        Returns:
            None
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyl
(Content truncated due to size limit. Use line ranges to read in chunks)